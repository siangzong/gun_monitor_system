import sys
from PyQt5.QtWidgets import QApplication, QLabel, QScrollArea, QVBoxLayout, QWidget
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
import os
from PyQt5 import QtWidgets
import sqlite3
from arduino_return import read_sql,process_data
import openpyxl

class MyWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("database_manage")
        base_dir = os.path.dirname(os.path.abspath(__file__))
        dir = os.path.join(base_dir, f"static/tk.ico")
        self.setWindowIcon(QIcon(dir))
        self.resize(800, 600)

        self.label = QLabel()
        self.label.setText('')
        self.label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.label.setGeometry(0, 0, 400, 600)
        self.label.setWordWrap(True)
        
        scroll_area = QScrollArea()
        scroll_area.setWidget(self.label)
        scroll_area.setMaximumSize(400, 600)

        layout = QVBoxLayout()
        layout.addWidget(scroll_area)
        self.setLayout(layout)

        self.s_btn = QtWidgets.QPushButton(self)
        self.s_btn.setText('查詢')
        self.s_btn.setGeometry(450, 50, 150, 80)
        self.s_btn.clicked.connect(self.select)
        self.i_btn = QtWidgets.QPushButton(self)
        self.i_btn.setText('新增')
        self.i_btn.setGeometry(600, 50, 150, 80)
        self.i_btn.clicked.connect(self.insert)
        self.u_btn = QtWidgets.QPushButton(self)
        self.u_btn.setText('修改')
        self.u_btn.setGeometry(450, 130, 150, 80)
        self.u_btn.clicked.connect(self.update)
        self.d_btn = QtWidgets.QPushButton(self)
        self.d_btn.setText('刪除')
        self.d_btn.setGeometry(600, 130, 150, 80)
        self.d_btn.clicked.connect(self.delete)
        self.output = QtWidgets.QPushButton(self)
        self.output.setText('匯出')
        self.output.setGeometry(450, 210, 150, 40)
        self.output.clicked.connect(self.make_excel)
        self.input = QtWidgets.QPushButton(self)
        self.input.setText('匯入')
        self.input.setGeometry(450, 250, 150, 40)
        self.input.clicked.connect(self.read_excel)


        self.output_time = QtWidgets.QLineEdit(self)   
        self.output_time.setGeometry(600, 210, 150, 40)
        self.output_time.setPlaceholderText("輸入匯出時間")
        self.input_time = QtWidgets.QLineEdit(self)   
        self.input_time.setGeometry(600, 250, 150, 40)
        self.input_time.setPlaceholderText("輸入匯入文件名稱")

        self.gun_id = QtWidgets.QLineEdit(self)   
        self.gun_id.setGeometry(600, 300, 150, 60)
        self.gun_id.setPlaceholderText("輸入槍枝序號")
        self.cb_gun_id = QtWidgets.QCheckBox(self)   
        self.cb_gun_id.move(750, 300)

        self.user_id = QtWidgets.QLineEdit(self)   
        self.user_id.setGeometry(450, 360, 150, 60)
        self.user_id.setPlaceholderText("輸入人員card_id")
        self.cb_user_id = QtWidgets.QCheckBox(self)   
        self.cb_user_id.move(430, 360)

        self.grade = QtWidgets.QLineEdit(self)   
        self.grade.setGeometry(600, 360, 150, 60)
        self.grade.setPlaceholderText("輸入級職")
        self.cb_grade = QtWidgets.QCheckBox(self)   
        self.cb_grade.move(750, 360)

        self.name = QtWidgets.QLineEdit(self)   
        self.name.setGeometry(450, 420, 150, 60)
        self.name.setPlaceholderText("輸入人員姓名")
        self.cb_name = QtWidgets.QCheckBox(self)   
        self.cb_name.move(430, 420)

        self.password = QtWidgets.QLineEdit(self)   
        self.password.setGeometry(600, 420, 150, 60)
        self.password.setPlaceholderText("輸入人員密碼")
        self.cb_password = QtWidgets.QCheckBox(self)   
        self.cb_password.move(750, 420)

        self.box = QtWidgets.QComboBox(self)  
        self.box.addItems(["entry_and_exit","cabinet","gun_state"])   
        self.box.setGeometry(450, 300, 150, 60)

       


    def select(self):
        gun_input = self.gun_id.text()
        user_id_input = self.user_id.text()
        grade_input = self.grade.text()
        name_input = self.name.text()
        pass_input = self.password.text()
        choose = self.box.currentText()
        con = sqlite3.connect("db.sqlite3")
        cur = con.cursor()
        try:

            if choose == "entry_and_exit":
                title="id,user_id,grade_level,name,state,time"
                query = "SELECT * FROM entry_and_exit WHERE 1=1"
                payload = []
                if user_id_input:
                    query += " AND user_id = ?"
                    payload.append(user_id_input)
                if grade_input:
                    query += " AND grade_level = ?"
                    payload.append(grade_input)
                if name_input:
                    query += " AND name = ?"
                    payload.append(name_input)
                cur.execute(query, payload)
                ans = cur.fetchall()
                final=[]    
                if ans:
                    for result in ans:
                        final.append(", ".join(map(str, result)))
                        formatted_results = "\n".join(final)
                        self.label.setText(f"{title}\n{formatted_results}")
            elif choose == "cabinet":
                title="id,grade_level,name,pass,state,time"
                query = "SELECT * FROM cabinet WHERE 1=1"
                payload = []
                if grade_input:
                    query += " AND grade_level = ?"
                    payload.append(grade_input)
                if name_input:
                    query += " AND name = ?"
                    payload.append(name_input)
                if pass_input:
                    query += " AND pass = ?"
                    payload.append(pass_input)
                cur.execute(query, payload)
                ans = cur.fetchall()
                final=[]    
                if ans:
                    for result in ans:
                        final.append(", ".join(map(str, result)))
                        formatted_results = "\n".join(final)
                        self.label.setText(f"{title}\n{formatted_results}")
            elif choose == "gun_state":
                title="id,gun_id,name,state,time"
                query = "SELECT * FROM gun_state WHERE 1=1"
                payload = []
                if gun_input:
                    query += " AND gun_id = ?"
                    payload.append(gun_input)
                if name_input:
                    query += " AND name = ?"
                    payload.append(name_input)
                cur.execute(query, payload)
                ans = cur.fetchall()
                final=[]    
                if ans:
                    for result in ans:
                        final.append(", ".join(map(str, result)))
                        formatted_results = "\n".join(final)
                        self.label.setText(f"{title}\n{formatted_results}")


            else:
                self.label.setText("無效的選擇")
                return
            
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
        finally:
            con.close()
      
    def insert(self):

        gun_input = self.gun_id.text()
        user_id_input = self.user_id.text()
        grade_input = self.grade.text()
        name_input = self.name.text()
        pass_input = self.password.text()
        choose = self.box.currentText()
        con = sqlite3.connect("db.sqlite3")
        cur = con.cursor()

        if not any([gun_input, user_id_input, grade_input, name_input, pass_input]):
            self.label.setText("未提供任何參數")
            return


        valid_tables = {"entry_and_exit", "cabinet", "gun_state"}

        if choose not in valid_tables:
            self.label.setText("無效的選擇")
            return

        try:
            if choose == "entry_and_exit":
                title = "id,user_id,grade_level,name,state,time"
                query = f"INSERT INTO {choose} (user_id, grade_level, name, state) VALUES (?, ?, ?, ?);"
                cur.execute(query, (user_id_input, grade_input, name_input, "exit"))
                con.commit()

                cur.execute(f"SELECT * FROM {choose} WHERE user_id = ?", (user_id_input,))
                ans = cur.fetchall()
                self.label.setText(f"{title}\n{ans}")

            elif choose == "cabinet":
                title = "id,grade_level,name,pass,state,time"
                query = f"INSERT INTO {choose} (grade_level, name, pass, state) VALUES (?, ?, ?, ?);"
                cur.execute(query, (grade_input, name_input, pass_input, "close"))
                con.commit()

                cur.execute(f"SELECT * FROM {choose} WHERE name = ?", (name_input,))
                ans = cur.fetchall()
                self.label.setText(f"{title}\n{ans}")

            elif choose == "gun_state":
                title = "id,gun_id,name,state,time"
                query = f"INSERT INTO {choose} (gun_id, name, state) VALUES (?, ?, ?);"
                cur.execute(query, (gun_input, name_input, "in"))
                con.commit()

                cur.execute(f"SELECT * FROM {choose} WHERE name = ?", (name_input,))
                ans = cur.fetchall()
                self.label.setText(f"{title}\n{ans}")

        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
            self.label.setText(f"插入失敗：{e}")
        finally:
            con.close()

    def update(self):
        gun_input = self.gun_id.text()
        ch_gun = self.cb_gun_id.isChecked()
        user_id_input = self.user_id.text()
        ch_user = self.cb_user_id.isChecked()
        grade_input = self.grade.text()
        ch_grade = self.cb_grade.isChecked()
        name_input = self.name.text()
        ch_name = self.cb_name.isChecked()
        pass_input = self.password.text()
        ch_pass = self.cb_password.isChecked()
        choose = self.box.currentText()

        con = sqlite3.connect("db.sqlite3")
        cur = con.cursor()

        if not any([gun_input, user_id_input, grade_input, name_input, pass_input]):
            self.label.setText("未提供任何參數")
            return

        try:
            if choose == "entry_and_exit":
                title = "id,user_id,grade_level,name,state,time"
                query = "UPDATE entry_and_exit SET "
                query_set = []
                query_where = " WHERE 1=1"
                payload = []

                if not ch_user and user_id_input:
                    query_set.append("user_id = ?")
                    payload.append(user_id_input)
                elif ch_user and user_id_input:
                    query_where += " AND user_id = ?"
                    payload.append(user_id_input)

                if not ch_grade and grade_input:
                    query_set.append("grade_level = ?")
                    payload.append(grade_input)
                elif ch_grade and grade_input:
                    query_where += " AND grade_level = ?"
                    payload.append(grade_input)

                if not ch_name and name_input:
                    query_set.append("name = ?")
                    payload.append(name_input)
                elif ch_name and name_input:
                    query_where += " AND name = ?"
                    payload.append(name_input)

                if not query_set:
                    self.label.setText("未指定要更新的欄位")
                    return

                query += ", ".join(query_set) + query_where
                cur.execute(query, payload)
                con.commit()

                cur.execute("SELECT * FROM entry_and_exit WHERE name = ?", (name_input,))
                ans = cur.fetchall()
                if ans:
                    self.label.setText(f"{title}\n{ans}")
                else:
                    self.label.setText("無結果符合更新條件")

            elif choose == "cabinet":
                title = "id,grade_level,name,pass,state,time"
                query = "UPDATE cabinet SET "
                query_set = []
                query_where = " WHERE 1=1"
                payload = []

                if not ch_grade and grade_input:
                    query_set.append("grade_level = ?")
                    payload.append(grade_input)
                elif ch_grade and grade_input:
                    query_where += " AND grade_level = ?"
                    payload.append(grade_input)

                if not ch_name and name_input:
                    query_set.append("name = ?")
                    payload.append(name_input)
                elif ch_name and name_input:
                    query_where += " AND name = ?"
                    payload.append(name_input)

                if not ch_pass and pass_input:
                    query_set.append("pass = ?")
                    payload.append(pass_input)
                elif ch_pass and pass_input:
                    query_where += " AND pass = ?"
                    payload.append(pass_input)

                if not query_set:
                    self.label.setText("未指定要更新的欄位")
                    return

                query += ", ".join(query_set) + query_where
                cur.execute(query, payload)
                con.commit()

                cur.execute("SELECT * FROM cabinet WHERE name = ?", (name_input,))
                ans = cur.fetchall()
                if ans:
                    self.label.setText(f"{title}\n{ans}")
                else:
                    self.label.setText("無結果符合更新條件")

            elif choose == "gun_state":
                title = "id,gun_id,name,state,time"
                query = "UPDATE gun_state SET "
                query_set = []
                query_where = " WHERE 1=1"
                payload = []

                if not ch_gun and gun_input:
                    query_set.append("gun_id = ?")
                    payload.append(gun_input)
                elif ch_gun and gun_input:
                    query_where += " AND gun_id = ?"
                    payload.append(gun_input)

                if not ch_name and name_input:
                    query_set.append("name = ?")
                    payload.append(name_input)
                elif ch_name and name_input:
                    query_where += " AND name = ?"
                    payload.append(name_input)

                if not query_set:
                    self.label.setText("未指定要更新的欄位")
                    return

                query += ", ".join(query_set) + query_where
                cur.execute(query, payload)
                con.commit()

                cur.execute("SELECT * FROM gun_state WHERE name = ?", (name_input,))
                ans = cur.fetchall()
                if ans:
                    self.label.setText(f"{title}\n{ans}")
                else:
                    self.label.setText("無結果符合更新條件")

            else:
                self.label.setText("無效的選擇")
                return

        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
            self.label.setText(f"更新失敗：{e}")
        finally:
            con.close()


    def delete(self):  
        gun_input = self.gun_id.text()
        user_id_input = self.user_id.text()
        grade_input = self.grade.text()
        name_input = self.name.text()
        pass_input = self.password.text()
        choose = self.box.currentText()
        con = sqlite3.connect("db.sqlite3")
        cur = con.cursor()
        if not any([gun_input, user_id_input, grade_input, name_input, pass_input]):
            self.label.setText("未提供任何參數")
            return

        try:
            if choose == "entry_and_exit":
                title="id,user_id,grade_level,name,state,time"
                query = "DELETE FROM entry_and_exit WHERE"
                payload = []
                if user_id_input:
                    query += "  user_id = ?"
                    payload.append(user_id_input)
                elif grade_input:
                    query += "  grade_level = ?"
                    payload.append(grade_input)
                elif name_input:
                    query += "  name = ?"
                    payload.append(name_input)
                cur.execute(query, payload)
                con.commit()
                cur.execute(f"刪除成功")


            elif choose == "cabinet":
                title="id,grade_level,name,pass,state,time"
                query = "DELETE FROM cabinet WHERE "
                payload = []
                if grade_input:
                    query += " grade_level = ?"
                    payload.append(grade_input)
                elif name_input:
                    query += " name = ?"
                    payload.append(name_input)
                elif pass_input:
                    query += " pass = ?"
                    payload.append(pass_input)
                cur.execute(query, payload)
                con.commit()
                cur.execute(f"刪除成功")


            elif choose == "gun_state":
                title="id,gun_id,name,state,time"
                query = "DELETE FROM gun_state WHERE "
                payload = []
                if gun_input:
                    query += "  gun_id = ?"
                    payload.append(gun_input)
                elif name_input:
                    query += "  name = ?"
                    payload.append(name_input)
                cur.execute(query, payload)
                con.commit()
                cur.execute(f"刪除成功")


            else:
                self.label.setText("無效的選擇")
                return

        
            
        except sqlite3.Error as e:
            print(f"An error occurred: {e}")
        finally:
            con.close()
    
    def make_excel(self):
        initial_dir = os.getcwd()#避免第二次點擊抓取文件時路徑不正確，故抓取當前路徑並在最後回歸
        target = self.output_time.text()
        with open(f'static/gun_state/{target}.log', "r") as f:
            lines = f.readlines()

            count=0
            parsed_data = []
            for line in lines:
                grade_level='student'
                reason='操課'
                parts = line.strip().split(",", 3)  
                name = parts[2] 
                gun_id = parts[1]
                act_time = parts[0]
                info=parts[3]
                parsed_data.append([grade_level,name,gun_id,reason,act_time,info])

            ans=[]
            id = 0
            while id < len(parsed_data):
                value = parsed_data[id]
                if value[5] == 'in':
                    in_time = value[4]
                    current_name = value[1]
                    ans.append([current_name, in_time])
                    parsed_data.pop(id)  
                else:
                    id += 1 

            i,k=0,0
            result=[]
            while len(ans)!=0:
                check=ans[i][0]
                k=0
                while len(parsed_data)!=0:
                    if parsed_data[k][1]==check:
                        final=parsed_data.pop(k)
                        final.insert(5,ans[i][1])
                        result.append(final)
                        break
                    else:
                        k+=1
                ans.pop(i)

            for i in range(len(result)):
                result[i].insert(0,i+1)
                result[i].pop(7)

            os.chdir('static/excel')  
            wb = openpyxl.load_workbook('get_and_use.xlsx', data_only=True)
            s1 = wb['工作表1'] 

            for i in range(len(result)):
                s1.cell(i+10,1).value =result[i][0]
                s1.cell(i+10,2).value =result[i][1]
                s1.cell(i+10,3).value =result[i][2]
                s1.cell(i+10,4).value =result[i][3]
                s1.cell(i+10,5).value =result[i][4]
                s1.cell(i+10,6).value =result[i][5]
                s1.cell(i+10,7).value =result[i][6]

            wb.save(f'{target}.xlsx')
        os.chdir(initial_dir)


    def read_excel(self):
        initial_dir = os.getcwd()
        target = self.input_time.text()
        choose = self.box.currentText()
        con = sqlite3.connect("db.sqlite3")
        cur = con.cursor()
        os.chdir('static/excel')

        wb = openpyxl.load_workbook(f'{target}.xlsx', data_only=True)  
        s1 = wb['工作表1']

        c = 0
        while s1.cell(c+3, 1).value is not None:
            c += 1

        data = []
        if choose == "entry_and_exit":
            title = "id, user_id, grade_level, name, state, time"
            for i in range(c):
                user_id = s1.cell(i+3, 1).value
                grade = s1.cell(i+3, 2).value
                name = s1.cell(i+3, 3).value
                data.append((user_id, grade, name, "exit")) 
            query = f"INSERT INTO {choose} (user_id, grade_level, name, state) VALUES (?, ?, ?, ?);"

        elif choose == "cabinet":
            title = "id, grade_level, name, pass, state, time"
            for i in range(c):
                grade = s1.cell(i+3, 1).value
                name = s1.cell(i+3, 2).value
                password = s1.cell(i+3, 3).value
                data.append((grade, name, password, "close")) 
            query = f"INSERT INTO {choose} (grade_level, name, pass, state) VALUES (?, ?, ?, ?);"

        elif choose == "gun_state":
            title = "id, gun_id, name, state, time"
            for i in range(c):
                gun_id = s1.cell(i+3, 1).value
                name = s1.cell(i+3, 2).value
                data.append((gun_id, name, "in"))  
            query = f"INSERT INTO {choose} (gun_id, name, state) VALUES (?, ?, ?);"

        else:
            self.label.setText("選擇的資料表無效")
            return

        try:
            cur.executemany(query, data)
            con.commit()
            cur.execute(f"SELECT * FROM {choose}")
            ans = cur.fetchall()
            self.label.setText(f"{title}\n{ans}")
        except Exception as e:
            self.label.setText(f"資料插入失敗：{e}")
        finally:
            con.close()
        os.chdir(initial_dir)


        

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWidget()  
    window.show()
    sys.exit(app.exec())
