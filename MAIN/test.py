from auto_sent import *
import sqlite3
from Lineapp.func import *  
import openpyxl
import os


os.chdir('static/excel')  # Colab 換路徑使用

import openpyxl
wb = openpyxl.load_workbook('test.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

s1 = wb['工作表1']

check=True
i=0
while check:
    if s1.cell(i+3,1).value==None:
        check=False
    else:
        i+=1
print(s1.cell(3,2).value)
print(s1.cell(4,2).value)
